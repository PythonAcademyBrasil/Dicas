from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

# Cria a Planilha Excel
arquivo_excel = Workbook()
planilha = arquivo_excel.active
planilha.title = "Dados"

# Adiciona Cabeçalho (primeira linha)
planilha.append(['Email', 'Valor Gasto', 'Qtde Itens', 'Valor médio por Item'])

# Trata dados de entrada do arquivo CSV
with open('./data/mock_data.csv') as csv:
    dados_arquivo = [linha.replace('\n', '').split(',') for linha in csv.readlines()]

# Itera sobre os dados do arquivo CSV, adicionando-os à Planilha
for dados in dados_arquivo:
    valor_gasto = float(dados[1])
    qtde_itens = int(dados[2])
    dados.append(f'{valor_gasto/qtde_itens:.2f}')
    dados[1] = valor_gasto
    dados[2] = qtde_itens

    planilha.append(dados)

# Mescla células
planilha.merge_cells('E1:F1')

# Cria Header de Resumo
header_resumo = planilha.cell(row=1, column=5)
header_resumo.value = 'RESUMO PROCESSAMENTO'
header_resumo.font = Font(name='Calibri', bold=True, color='FFFFFF')
header_resumo.alignment = Alignment(horizontal='center', vertical='center')
header_resumo.fill = PatternFill('solid', '222222')

# Cria Subheader de Resumo do Valor Total
subheader_valor = planilha.cell(row=2, column=5)
subheader_valor.value = 'VALOR TOTAL'
subheader_valor.fill = PatternFill('solid', '999999')

# Cria Subheader de Resumo da Quantidade Total
subheader_qtde = planilha.cell(row=2, column=6)
subheader_qtde.value = 'QUANTIDADE TOTAL DE ITENS'
subheader_qtde.fill = PatternFill('solid', '999999')

# Cria Célula com Fórmula formatado para Moeda
soma_valor = planilha.cell(row=3, column=5)
soma_valor.value = '=SUM(B2:B251)'
soma_valor.number_format = '[$R$-416]\\ #,##0.00'
soma_valor.fill = PatternFill('solid', 'DDDDDD')

# Cria Célula com Fórmula
soma_qtde = planilha.cell(row=3, column=6)
soma_qtde.value = '=SUM(C2:C251)'
soma_qtde.fill = PatternFill('solid', 'DDDDDD')

# Salva o arquivo XLSX de saída
arquivo_excel.save("relatorio.xlsx")
